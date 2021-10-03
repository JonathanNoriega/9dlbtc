from openpyxl.utils import get_column_letter
from py_currency_converter import convert
from infi.systray import SysTrayIcon
from openpyxl import load_workbook
from binance.client import Client
from openpyxl import Workbook
from time import strftime
import time
import os
import pandas as pd
import matplotlib.pyplot as plt







api_key = os.environ.get('binance_api')
api_secret = os.environ.get('binance_secret')

#btc price

amount = 0.0001273

client = Client(api_key, api_secret)
btc_price = client.get_symbol_ticker(symbol="BTCUSDT")
usd_price = float(btc_price["price"])

price = convert(base='USD', amount=usd_price, to=['COP'])
print(price['COP'])     

#Excel

book = load_workbook("log.xlsx")
sheet = book.active
 


def bye(systray):
    os.system("taskkill /f /im  python.exe")

def graph(systray):
    data = pd.read_excel('C:\\Users\\User\\Desktop\\Coding\\Python\\9dbtc\\log.xlsx')
    precio = data['worth']
    data.head()
    plt.plot( precio, color='green')
    plt.title('My 9$ bitcoin investment', fontsize=14)
    plt.ylabel('9$ worth of bitcoin', fontsize=14)
    plt.grid(True)
    plt.show()
    
menu_options = (("Show Graph", None, graph),)
systray = SysTrayIcon("../9dbtc/img/btc.ico", "9dbtc", menu_options, on_quit=bye)
systray.start()



def save_data(delay):
    sheet.column_dimensions["A"].width = 25
    sheet.column_dimensions["B"].width = 25
    sheet.column_dimensions["C"].width = 25
    sheet.column_dimensions["D"].width = 25
    while True:
        client = Client(api_key, api_secret)
        btc_price = client.get_symbol_ticker(symbol="BTCUSDT")
        usd_price = float(btc_price["price"])
        price = convert(base='USD', amount=usd_price, to=['COP'])
        d9 = amount * float(price['COP']) / 1
        
        sheet['A1'] = "Date"
        sheet['B1'] = "price"
        sheet['C1'] = "worth"
        lastRowA = 'A' + str(sheet.max_row + 1)
        lastRowB = 'B' + str(sheet.max_row + 1)
        lastRowC = 'C' + str(sheet.max_row + 1)
        sheet[lastRowA] = strftime("%a, %d %b %Y %H:%M:%S")
        sheet[lastRowB] = price['COP']
        sheet[lastRowC] = d9
        book.save('log.xlsx')
        time.sleep(delay)






print("#####################################")
print("#                                   #")
print("#                                   #")
print("#     Saving data in 'log.xlsx'     #")
print("#                                   #")
print("#                                   #")
print("#####################################")

# save_data(1800)



